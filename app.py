import os
import logging
from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash, make_response
from werkzeug.utils import secure_filename
import openpyxl
import google.generativeai as genai
import markdown
from docx import Document
from io import BytesIO

# Import for password protection
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'your_default_secret_key')

# --- Configuration ---
API_KEY = os.getenv("GEMINI_API_KEY")
DEFAULT_MODEL_NAME = 'gemini-2.0-flash'
THINKING_MODEL_NAME = 'gemini-2.0-flash-thinking-exp'
SYSTEM_PROMPT = """You are an expert in analyzing Excel spreadsheets and explaining their purpose, structure, and formulas in a clear and concise manner.
Your goal is to provide a detailed explanation of the spreadsheet's content, including the meaning of each column, the data structure, and the purpose of any formulas.
You should format your explanations in Markdown, using headings, bullet points, and code blocks where appropriate to enhance readability.
Focus on providing insights that would be helpful to someone unfamiliar with the spreadsheet."""
FORMULA_SYSTEM_PROMPT = """You are an expert Excel formula creator. I will describe what I need a formula for, and you will provide the most efficient and correct Excel formula to achieve this.
You should explain the formula and provide an example of how to use it. If there are multiple ways to achieve it, provide the best and most common approach unless specified otherwise.
Assume I am using standard Excel functions."""
RECONCILIATION_SYSTEM_PROMPT = """You are an expert in accounts reconciliation using Excel spreadsheets.
Your task is to compare two Excel sheets, identify discrepancies, and explain the differences clearly and concisely in Markdown format.

Consider the two sheets as representing two sets of accounts or data for reconciliation.
Identify and explain:
- Any differences in data between corresponding columns or rows in the two sheets.
- Any missing entries or rows that are present in one sheet but not in the other.
- Any inconsistencies in formatting or data types that might indicate discrepancies.
- For numerical data, highlight any significant differences or variances.

Structure your reconciliation report in Markdown with:
- Clear headings for each type of discrepancy found.
- Bullet points listing specific discrepancies.
- Code blocks or tables where appropriate to show data differences or examples.
- A summary of the overall reconciliation status and key findings.

Focus on providing a report that is easy to understand for someone needing to reconcile these accounts and identify potential issues."""


PROMPT_PREFIX = "The Excel sheet contains the following information in a structured way:\n"
PROMPT_SUFFIX = "\nPlease provide a detailed explanation in Markdown format. Explain what this sheet is about, what each column/section represents, and how the data is structured and what its purpose might be. If there are formulas, explain their logic in simple terms. Structure your answer with headings, bullet points, and code blocks where appropriate for formulas or data examples to enhance readability."
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
DEFAULT_DOCX_FILENAME = "excel_explanation.docx"
FORMULA_DOCX_FILENAME = "excel_formula.docx"
CHAT_DOCX_FILENAME = "excel_chat.docx"
RECONCILIATION_DOCX_FILENAME = "excel_reconciliation.docx"


os.makedirs(UPLOAD_FOLDER, exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Flask-Login Configuration ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # Define the login view function
login_manager.login_message = "Please log in to access this page." # Custom message

# --- User Data from Environment Variables ---
users = {}
for i in range(1, 5):  # Load data for 4 users (User1, User2, User3, User4)
    username = os.getenv(f"USER{i}_USERNAME")
    password = os.getenv(f"USER{i}_PASSWORD")
    if username and password: # Only add user if both username and password env vars are set
        users[i] = {'username': username, 'password_hash': generate_password_hash(password)}
    else:
        logging.warning(f"User {i} credentials not fully configured via environment variables (USER{i}_USERNAME, USER{i}_PASSWORD). User {i} will not be available for login.")


class User(UserMixin):
    def __init__(self, id, username, password_hash):
        self.id = id
        self.username = username
        self.password_hash = password_hash

    def verify_password(self, password):
        return check_password_hash(self.password_hash, password)

@login_manager.user_loader
def load_user(user_id):
    user_data = users.get(int(user_id))
    if user_data:
        return User(id=user_id, username=user_data['username'], password_hash=user_data['password_hash'])
    return None


def configure_api():
    """Configures the Gemini API with the API key."""
    if not API_KEY:
        logging.error("API_KEY environment variable not set.")
        return False
    try:
        genai.configure(api_key=API_KEY)
        return True
    except Exception as e:
        logging.error(f"Error configuring Gemini API: {e}")
        return False


def allowed_file(filename):
    """Check if the uploaded file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def load_excel_data(file_path):
    """Loads data from the Excel file using openpyxl."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        sheet = wb.active
        return sheet
    except Exception as e:
        logging.error(f"Error loading Excel file: {e}")
        return None

def build_prompt_reconciliation(sheet1, sheet2):
    """Builds the prompt for Gemini API for reconciliation, comparing two sheets."""
    prompt_content = "Sheet 1 Data:\n"
    for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
        row_values = []
        for cell in row:
            row_values.append(str(cell.value) if cell.value is not None else "None")
        prompt_content += "- Row: " + ", ".join(row_values) + "\n"

    prompt_content += "\nSheet 2 Data:\n"
    for row in sheet2.iter_rows(min_row=1, max_row=sheet2.max_row, min_col=1, max_col=sheet2.max_column):
        row_values = []
        for cell in row:
            row_values.append(str(cell.value) if cell.value is not None else "None")
        prompt_content += "- Row: " + ", ".join(row_values) + "\n"

    full_prompt = RECONCILIATION_SYSTEM_PROMPT + "\n\nData from Sheet 1 and Sheet 2 to reconcile:\n" + prompt_content
    logging.info("Reconciliation prompt built successfully.")
    return full_prompt


def build_prompt(sheet):
    """Builds the prompt for the Gemini API based on the Excel sheet data."""
    prompt_content = ""
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None or cell.comment is not None:
                cell_info = ""
                if cell.data_type == 'f':
                    cell_info = f"formula '{cell.value}'"
                elif cell.value is not None:
                    cell_info = f"value '{cell.value}'"
                else:
                    cell_info = "no value"

                comment_text = ""
                if cell.comment:
                    comment_text_raw = cell.comment.text.strip()
                    comment_text_processed = comment_text_raw.replace('\n', ' ') # Process newline *before* f-string
                    comment_text = f" with comment '{comment_text_processed}'"

                prompt_content += f"- Cell {cell.coordinate} has {cell_info}{comment_text}.\n"

    full_prompt = PROMPT_PREFIX + prompt_content + PROMPT_SUFFIX
    logging.info("Prompt built successfully.")
    return full_prompt


def get_explanation_from_gemini(prompt, model_name):
    """Gets explanation from Gemini API."""
    model = genai.GenerativeModel(model_name)
    try:
        response = model.generate_content(prompt, generation_config=genai.types.GenerationConfig(temperature=0.2)) # slight temp increase for reconciliation & formula
        explanation = response.text
        logging.info(f"Explanation received from Gemini API using model: {model_name}")
        return explanation
    except Exception as e:
        logging.error(f"Error communicating with Gemini API: {e}")
        return None

def get_formula_from_gemini(prompt):
    """Gets formula from Gemini API using formula system prompt."""
    model = genai.GenerativeModel(DEFAULT_MODEL_NAME) # Using default model for formula gen
    full_prompt = FORMULA_SYSTEM_PROMPT + "\n\n" + prompt
    try:
        response = model.generate_content(full_prompt, generation_config=genai.types.GenerationConfig(temperature=0.4)) #added temperature
        formula_explanation = response.text
        logging.info("Formula explanation received from Gemini API.")
        return formula_explanation
    except Exception as e:
        logging.error(f"Error communicating with Gemini API for formula: {e}")
        return None


def export_to_docx(explanation, filename=DEFAULT_DOCX_FILENAME):
    """Exports the explanation to a DOCX file in memory and returns BytesIO object."""
    doc = Document()
    for line in explanation.splitlines():
        doc.add_paragraph(line)

    docx_stream = BytesIO()
    try:
        doc.save(docx_stream)
        docx_stream.seek(0)
        logging.info(f"Explanation exported to DOCX in memory as {filename}.")
        return docx_stream
    except Exception as e:
        logging.error(f"Error exporting to DOCX: {e}")
        return None


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page."""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user_data = None
        for user_id, data in users.items(): # Simple user lookup from env vars
            if data['username'] == username:
                user_data = data
                user_id_found = user_id
                break

        if user_data and check_password_hash(user_data['password_hash'], password):
            user = User(id=user_id_found, username=username, password_hash=user_data['password_hash'])
            login_user(user)
            flash('Logged in successfully.')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index')) # Redirect to original page or index
        else:
            flash('Invalid username or password', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully.')
    return redirect(url_for('index'))


@app.route('/', methods=['GET', 'POST'])
@login_required # Protect the index page
def index():
    """Handles the main application logic for Excel sheet explanation."""
    explanation_html = None
    docx_stream = None
    error = None
    model_name = DEFAULT_MODEL_NAME

    if request.method == 'POST':
        if 'excel_file' not in request.files:
            error = 'No file part'
        elif request.files['excel_file'].filename == '':
            error = 'No selected file'
        elif 'excel_file' in request.files and allowed_file(request.files['excel_file'].filename): # check file is in request.files
            file = request.files['excel_file']
            try:
                filename = secure_filename(file.filename)
                file_path = os.path.join(UPLOAD_FOLDER, filename)
                file.save(file_path)

                selected_model = request.form.get('model_select')
                if selected_model == 'thinking':
                    model_name = THINKING_MODEL_NAME
                else:
                    model_name = DEFAULT_MODEL_NAME

                sheet = load_excel_data(file_path)
                if sheet:
                    prompt = build_prompt(sheet)
                    explanation_markdown = get_explanation_from_gemini(prompt, model_name)

                    if explanation_markdown:
                        explanation_html = markdown.markdown(explanation_markdown)
                        session['explanation_markdown'] = explanation_markdown
                        session['current_explanation_html'] = explanation_html # Store html for chat context
                    else:
                        error = "Failed to get explanation from Gemini API."
                else:
                    error = "Failed to load Excel data."
            except Exception as e:
                error = f"An error occurred: {e}"
            finally:
                os.remove(file_path)
        else:
            error = 'Invalid file type. Allowed types are xlsx, xls'

    response = make_response(render_template('index.html', explanation_html=explanation_html, error=error, model_name=model_name, current_user=current_user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/export_docx')
@login_required # Protect export
def export_docx_route():
    """Exports the explanation to DOCX format and allows download."""
    explanation_markdown = session.get('explanation_markdown')
    if not explanation_markdown:
        return "No explanation available to export.", 400

    docx_stream = export_to_docx(explanation_markdown, DEFAULT_DOCX_FILENAME)
    if docx_stream:
        return send_file(
            docx_stream,
            as_attachment=True,
            download_name=DEFAULT_DOCX_FILENAME,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    else:
        return "Error exporting to DOCX.", 500

@app.route('/formula_creator', methods=['GET', 'POST'])
@login_required # Protect formula creator
def formula_creator():
    """Handles the formula creation page."""
    formula_explanation_html = None
    docx_stream = None
    error = None

    if request.method == 'POST':
        formula_description = request.form.get('formula_description')
        if formula_description:
            formula_explanation_markdown = get_formula_from_gemini(formula_description)
            if formula_explanation_markdown:
                formula_explanation_html = markdown.markdown(formula_explanation_markdown)
                session['formula_explanation_markdown'] = formula_explanation_markdown
            else:
                error = "Failed to get formula explanation from Gemini API."
        else:
            error = "Please enter a description for the formula you need."

    response = make_response(render_template('formula_creator.html', formula_explanation_html=formula_explanation_html, error=error, current_user=current_user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/export_formula_docx')
@login_required # Protect formula export
def export_formula_docx_route():
    """Exports the formula explanation to DOCX format."""
    formula_explanation_markdown = session.get('formula_explanation_markdown')
    if not formula_explanation_markdown:
        return "No formula explanation available to export.", 400

    docx_stream = export_to_docx(formula_explanation_markdown, FORMULA_DOCX_FILENAME)
    if docx_stream:
        return send_file(
            docx_stream,
            as_attachment=True,
            download_name=FORMULA_DOCX_FILENAME,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    else:
        return "Error exporting formula explanation to DOCX.", 500

@app.route('/chat', methods=['GET', 'POST'])
@login_required # Protect chat page
def chat():
    """Handles the chat functionality after sheet analysis."""
    explanation_html = session.get('current_explanation_html') # Get html explanation for display
    chat_history = session.get('chat_history', [])
    user_message = None
    error = None
    docx_stream = None

    if explanation_html is None: #Redirect if no explanation is available
        return redirect(url_for('index'))

    if request.method == 'POST':
        user_message = request.form.get('chat_message')
        if user_message:
            prompt_context = f"The analysis of the Excel sheet is:\n\n{session.get('explanation_markdown')}\n\nUser's question: {user_message}"
            llm_response_markdown = get_explanation_from_gemini(prompt_context, DEFAULT_MODEL_NAME) # Or thinking model?
            if llm_response_markdown:
                llm_response_html = markdown.markdown(llm_response_markdown)
                chat_history.append({'user': user_message, 'bot': llm_response_html})
                session['chat_history'] = chat_history
            else:
                error = "Failed to get chat response from Gemini API."
        else:
            error = "Please enter a chat message."

    response = make_response(render_template('chat.html', explanation_html=explanation_html, chat_history=chat_history, error=error, current_user=current_user))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/export_chat_docx')
@login_required # Protect chat export
def export_chat_docx_route():
    """Exports the chat history to DOCX format."""
    chat_history = session.get('chat_history')
    if not chat_history:
        return "No chat history available to export.", 400

    chat_markdown = ""
    for message in chat_history:
        chat_markdown += f"**User:** {message['user']}\n\n"
        chat_markdown += f"**Bot:** {message['bot']}\n\n---\n\n"

    docx_stream = export_to_docx(chat_markdown, CHAT_DOCX_FILENAME)
    if docx_stream:
        return send_file(
            docx_stream,
            as_attachment=True,
            download_name=CHAT_DOCX_FILENAME,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    else:
        return "Error exporting chat history to DOCX.", 500

@app.route('/reconcile', methods=['GET', 'POST'])
@login_required
def reconcile():
    """Handles the accounts reconciliation page and logic."""
    reconciliation_explanation_html = None
    docx_stream = None
    error = None
    model_name = DEFAULT_MODEL_NAME # Default model

    if request.method == 'POST':
        if 'excel_file_1' not in request.files or 'excel_file_2' not in request.files:
            error = 'Need to upload both Sheet 1 and Sheet 2'
        elif request.files['excel_file_1'].filename == '' or request.files['excel_file_2'].filename == '':
            error = 'Both Sheet 1 and Sheet 2 files need to be selected'
        elif 'excel_file_1' in request.files and allowed_file(request.files['excel_file_1'].filename) and 'excel_file_2' in request.files and allowed_file(request.files['excel_file_2'].filename):
            file1 = request.files['excel_file_1']
            file2 = request.files['excel_file_2']
            file_path_1 = os.path.join(UPLOAD_FOLDER, secure_filename(file1.filename))
            file_path_2 = os.path.join(UPLOAD_FOLDER, secure_filename(file2.filename))

            try:
                file1.save(file_path_1)
                file2.save(file_path_2)

                selected_model = request.form.get('model_select') # Get selected model
                if selected_model == 'thinking':
                    model_name = THINKING_MODEL_NAME
                else:
                    model_name = DEFAULT_MODEL_NAME


                sheet1 = load_excel_data(file_path_1)
                sheet2 = load_excel_data(file_path_2)

                if sheet1 and sheet2:
                    prompt = build_prompt_reconciliation(sheet1, sheet2)
                    reconciliation_markdown = get_explanation_from_gemini(prompt, model_name) # Use selected model

                    if reconciliation_markdown:
                        reconciliation_explanation_html = markdown.markdown(reconciliation_markdown)
                        session['reconciliation_explanation_markdown'] = reconciliation_markdown
                    else:
                        error = "Failed to get reconciliation explanation from Gemini API."
                else:
                    error = "Failed to load data from one or both Excel files."
            except Exception as e:
                error = f"An error occurred during reconciliation: {e}"
            finally:
                if os.path.exists(file_path_1):
                    os.remove(file_path_1)
                if os.path.exists(file_path_2):
                    os.remove(file_path_2)
        else:
            error = 'Invalid file types. Allowed types are xlsx, xls for both sheets.'

    response = make_response(render_template('reconcile.html', reconciliation_explanation_html=reconciliation_explanation_html, error=error, current_user=current_user, model_name=model_name)) # Pass model_name to template
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/export_reconciliation_docx')
@login_required
def export_reconciliation_docx_route():
    """Exports the reconciliation explanation to DOCX format."""
    reconciliation_explanation_markdown = session.get('reconciliation_explanation_markdown')
    if not reconciliation_explanation_markdown:
        return "No reconciliation explanation available to export.", 400

    docx_stream = export_to_docx(reconciliation_explanation_markdown, RECONCILIATION_DOCX_FILENAME)
    if docx_stream:
        return send_file(
            docx_stream,
            as_attachment=True,
            download_name=RECONCILIATION_DOCX_FILENAME,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    else:
        return "Error exporting reconciliation explanation to DOCX.", 500


if __name__ == '__main__':
    if configure_api():
        app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
